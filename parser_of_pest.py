from urllib.parse import urljoin
import requests,bs4
from time import sleep
import openpyxl

url_to_scrap='http://www.agriculture.gov.au/pests-diseases-weeds/plant#identify-pests-diseases'
base_url=urljoin(url_to_scrap,'/')[:-1]

def getme():
	print (base_url)
	res=requests.get(url_to_scrap)
	res.raise_for_status()
	soup=bs4.BeautifulSoup(res.text,"html.parser")
	no_of_items=soup.select('.flex-item')
	data=[]

	for i in range(len(no_of_items)):
		link_name=soup.select('.flex-item')[i].find('a').text[1:]
		linkElem=soup.select('.flex-item a')[i]
		link_next=linkElem.get('href')
		if not link_next.startswith('http'):
			link_next=base_url+link_next
		link_img=soup.select('.flex-item img')[i].get('src')
		link_img=base_url+link_img
		data.append([link_name,link_next,link_img])
	return data


def find_diseases(soup):
	"""the core diseas of pest  """
	diseases_list=[]
	try:
		diseases=soup.select('#collapsefaq ul')[0]
		for diseas in diseases:
			diseases_list.append(diseas.text)
			
	except:
		print("desired information not found")
	return diseases_list






def origin_of_pest(soup):
	"""return a string which is a origin of pest """
	html=''
	try:
		sec=soup.find_all("strong")[1]
		for tag in sec.next_siblings:
			if tag.name == "strong":
				break
			else:
				html+=str(tag)
		origin=html[:-5]
	except:
		print("unable to locate origin")
		origin='not available'

	return origin



def legally_to_aus(soup):
	"""Check what can legally come into Australia  """
	try:
		sec=soup.find_all('div',attrs={'class':'hide'})[1].find_all('li')
		come_in_to_aus=[]
		for i in sec:
			come_in_to_aus.append(i.text)
	except:
		come_in_to_aus=[]
	return come_in_to_aus



def suspect_specimen(soup):
	""" Secure any suspect specimens   """
	try:
		sec=soup.find_all('div',attrs={'class':'hide'})[2].find('p')
		
	except:
		print("no data found secure any suspect specimen")
		return 'not available'
	return sec.text




def identify(soup):
	""" see if you can identify  """
	try:
		asc=soup.find('div',attrs={'class':'hide'}).find_all('ul')
		identify=[]
		for i in asc:
			identify.append(i.getText())
		
	except:
		print("see if u can identify error ")
		identify=['not available']
	return identify

def string_clean(raw):
	neat=raw.lstrip(' ')
	
def write_header(sheet):
	"""to bulid the header for excel sheet file"""
	headers=['NAME','IMAGE LINK','ORIGIN','IDENTIFYING THE PEST','LEGALLY TO AUSTRALIA','SECURE SUSPECT SPECIMENS']
	for i in range(1,2):
		for k in range(1,7):
			sheet.cell(row=i, column=k).value = headers[k-1]
	for i in range(1,2):
		for k in range(1,7):
			sheet.cell(row=i, column=k).font = openpyxl.styles.Font(bold=True, italic=True)	


def getme_again(data):
	book=openpyxl.Workbook()
	#book.create_sheet('sample sheet')
	sheet=book.get_sheet_by_name('Sheet')
	write_header(sheet)
	new_data=[]
	#print(data)
	links=[(link[1]) for link in data  if  link[1].startswith('http://www.agriculture.gov.au')]
	sec=0
	rows=2
	print(len(links))
	while(sec<len(links)):
		link=links[sec]
		name=data[sec][0]
		link_img=data[sec][2]
		res=requests.get(link)
		res.raise_for_status()
		soup=bs4.BeautifulSoup(res.text,"lxml")
		origin=origin_of_pest(soup)
		ident=''.join(identify(soup))
		legal=''.join(legally_to_aus(soup))
		suspect=suspect_specimen(soup)
		lit=[name,link_img,origin,ident,legal,suspect]
		for k in range(1,7):
			sheet.cell(row=rows, column=k).value = lit[k-1]

		
		rows+=1
		sec+=1
		
	book.save('Scraped_data.xlsx')	


def main():
	getme_again(getme())


if __name__ == '__main__':
	main()
